from django.shortcuts import render, redirect
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from .forms import UserForm, NewsForm
from .models import User, News
from .decorators import authenticated_user, allowed_user

# Create your views here.
def home(request):
	return render(request, 'main/home.html', {})



def about_page(request):
	success = None
	students = User.objects.all()
	if request.method == 'POST':
		success = print_excel(request.POST)

	context = {
		"students":students,
		"success": success
	}
	return render(request, 'main/about.html', context)



def login_page(request):
	message =None
	if request.method == 'POST':
		email = request.POST.get('email')
		password = request.POST.get('password')
		try:
			user = User.objects.get(email=email)
		except:
			messages.error(request, "user does not exist")
			
		user = authenticate(request, email=email, password=password)
		if user is not None:
			login(request, user)
			return redirect('home')
		else:
			message = "Username or Password is incorrect!"
			messages.info(request, "Username or Password is incorrect!")

	context = {
	"messages":message
	}
	return render(request, 'register/login.html', context)

def logout_page(request):
	logout(request)
	return render(request, 'main/home.html', {})


def sign_up(request):
	if request.method == 'POST':
		form = UserForm(request.POST)
		if form.is_valid():
			user = form.save()
			login(request, user)
			return redirect('home')
	else:
		form = UserForm()
	context = {"form": form}
	return render(request, 'register/sign-up.html', context)

@allowed_user(allowed_group=['Student', 'Leader'])
def news_page(request):
	news = News.objects.all()
	if request.method == 'POST':
		form = NewsForm(request.POST)
		if form.is_valid():
			news = form.save(commit=False)
			news.added_by = request.user
			news.save()
			return redirect('news')
	else:
		form = NewsForm()
	context ={
	"news":news,
	"form":form
	}
	return render(request, 'main/news.html', context)

def restricted(request):
	return render(request, 'error/restricted.html', {})	





def print_excel(request):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font


    students = User.objects.all()
    wb = Workbook()
    w_sheet = wb.active
    no_students = students.count()
    w_sheet.title = "Students"

    if students:
        headings  = ["ID", "First Name", "Last Name", "Username", "Email", "Phone Number"]
        w_sheet.append(headings)
        # w_sheet[headings].font = Font(bold=True)
        for student in students:
            
            student_list = [student.id,student.first_name,  student.last_name, student.username, student.email, student.phone_number]

            for col in range(1, 7):
                char = get_column_letter(col)
                w_sheet[char + str(student.id+1)] = student_list[col-1]

        headings  = ["Total = ", f"{no_students}"]
        w_sheet.append(headings)
    for col in range(1, 7):
        w_sheet[get_column_letter(col)+ '1'].font = Font(bold=True, color="00FF2311")
    wb.save("C:/Users/USER/Desktop/student.xlsx")
    return "C:/Users/USER/Desktop/student.xlsx"




